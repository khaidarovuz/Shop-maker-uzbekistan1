"""
Yordamchi funksiyalar.
"""

import io
import json
import logging
from datetime import datetime, date
from typing import Any
import aiosqlite

logger = logging.getLogger(__name__)


def format_price(price: float) -> str:
    """Narxni formatlaydi: 10000.0 → '10 000 so'm'"""
    return f"{price:,.0f}".replace(",", " ") + " so'm"


def format_date(dt: datetime | str | None) -> str:
    """Sanani formatlaydi."""
    if not dt:
        return "—"
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt)
        except ValueError:
            return dt
    return dt.strftime("%d.%m.%Y %H:%M")


def format_date_short(dt: datetime | str | None) -> str:
    """Qisqa sana formati."""
    if not dt:
        return "—"
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt)
        except ValueError:
            return dt
    return dt.strftime("%d.%m.%Y")


def paginate(items: list, page: int, per_page: int = 10) -> tuple[list, int, int]:
    """
    Ro'yxatni sahifalaydi.
    Qaytaradi: (sahifa_elementlari, jami_sahifalar, joriy_sahifa)
    """
    total = len(items)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    end = start + per_page
    return items[start:end], total_pages, page


def get_plan_label(plan: str) -> str:
    """Reja nomini belgi bilan qaytaradi."""
    labels = {
        "free": "🆓 Bepul",
        "comfort": "🌟 Comfort",
        "premium": "💎 Premium",
    }
    return labels.get(plan, "🆓 Bepul")


def get_order_status_label(status: str) -> str:
    """Buyurtma holati nomini qaytaradi."""
    statuses = {
        "new": "🆕 Yangi",
        "processing": "⏳ Jarayonda",
        "shipped": "🚚 Yetkazilmoqda",
        "delivered": "✅ Yetkazildi",
        "cancelled": "❌ Bekor qilindi",
    }
    return statuses.get(status, status)


def order_items_summary(items_json: str) -> str:
    """Buyurtma mahsulotlari qisqa ko'rinishini qaytaradi."""
    try:
        items = json.loads(items_json)
        lines = []
        for item in items:
            lines.append(
                f"• {item['name']} × {item['qty']} = {format_price(item['price'] * item['qty'])}"
            )
        return "\n".join(lines)
    except Exception:
        return "Noma'lum mahsulotlar"


def build_order_text(order: aiosqlite.Row | dict, include_header: bool = True) -> str:
    """Buyurtma matnini qaytaradi."""
    if isinstance(order, aiosqlite.Row):
        order = dict(order)

    items_text = order_items_summary(order.get("items", "[]"))
    status = get_order_status_label(order.get("status", "new"))

    text = ""
    if include_header:
        text += f"🧾 <b>Buyurtma #{order['id']}</b>\n"
        text += f"📅 {format_date(order.get('created_at'))}\n\n"

    text += f"👤 Xaridor: {order.get('customer_name', '—')}\n"
    if order.get("customer_phone"):
        text += f"📞 Telefon: {order['customer_phone']}\n"
    if order.get("customer_address"):
        text += f"📍 Manzil: {order['customer_address']}\n"
    text += f"\n📦 Mahsulotlar:\n{items_text}\n"
    text += f"\n💰 Jami: <b>{format_price(order.get('total_price', 0))}</b>\n"
    text += f"📊 Holat: {status}"
    if order.get("note"):
        text += f"\n📝 Izoh: {order['note']}"

    return text


async def export_orders_to_excel(orders: list) -> io.BytesIO:
    """
    Buyurtmalarni Excel fayliga eksport qiladi.
    openpyxl ishlatadi.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Buyurtmalar"

        # Sarlavhalar
        headers = [
            "ID", "Xaridor", "Telefon", "Manzil",
            "Mahsulotlar", "Jami narx", "Holat", "Sana"
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fgColor="4472C4", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Ma'lumotlar
        for row_idx, order in enumerate(orders, 2):
            if isinstance(order, aiosqlite.Row):
                order = dict(order)

            items_summary = ""
            try:
                items = json.loads(order.get("items", "[]"))
                parts = [f"{i['name']} x{i['qty']}" for i in items]
                items_summary = "; ".join(parts)
            except Exception:
                pass

            ws.cell(row=row_idx, column=1, value=order.get("id"))
            ws.cell(row=row_idx, column=2, value=order.get("customer_name", ""))
            ws.cell(row=row_idx, column=3, value=order.get("customer_phone", ""))
            ws.cell(row=row_idx, column=4, value=order.get("customer_address", ""))
            ws.cell(row=row_idx, column=5, value=items_summary)
            ws.cell(row=row_idx, column=6, value=order.get("total_price", 0))
            ws.cell(row=row_idx, column=7, value=get_order_status_label(order.get("status", "")))
            ws.cell(row=row_idx, column=8, value=str(order.get("created_at", "")))

        # Ustun kengliklarini moslashtirish
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(max_length, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    except ImportError:
        logger.error("openpyxl o'rnatilmagan!")
        raise


def chunks(lst: list, n: int):
    """Ro'yxatni n o'lchamli qismlarga bo'ladi."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def mention_user(full_name: str, user_id: int) -> str:
    """Foydalanuvchiga HTML mention yaratadi."""
    return f'<a href="tg://user?id={user_id}">{full_name}</a>'
